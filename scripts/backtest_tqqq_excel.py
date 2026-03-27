"""
TQQQ EMA/SMA Strategy — Excel Report

Generates a multi-sheet Excel workbook with:
  1. Summary — all combos with metrics, color-coded
  2. Yearly Returns — every combo × year matrix
  3. Key Combos — detailed comparison
  4. Consistency — top 10 / bottom 5 rankings
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from itertools import product
from pathlib import Path

INITIAL_CAPITAL = 100_000
CASH_YIELD = 0.045
ENTRY_EMAS = sorted(set([5, 8, 10, 13, 15, 20, 25, 30] + [3, 7, 11, 17, 19, 23, 29]))
EXIT_SMAS = sorted(set([5, 10, 15, 20, 25, 30, 40, 50] + [3, 7, 11, 17, 19, 23, 29]))
REPORT_PATH = Path(__file__).parent.parent / "reports" / "tqqq_strategy_report.xlsx"

# Style constants
GREEN = "4CAF50"
RED = "EF5350"
DARK_BG = "1A1A2E"
HEADER_BG = "2A2A3E"
ORANGE = "FF9800"
BLUE = "64B5F6"
WHITE = "E0E0E0"
LIGHT_GREEN_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="4A1010", end_color="4A1010", fill_type="solid")
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
DARK_FILL = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
BENCHMARK_FILL = PatternFill(start_color="333355", end_color="333355", fill_type="solid")
HEADER_FONT = Font(name="Consolas", bold=True, color=ORANGE, size=11)
DATA_FONT = Font(name="Consolas", color=WHITE, size=10)
GREEN_FONT = Font(name="Consolas", color=GREEN, size=10)
RED_FONT = Font(name="Consolas", color=RED, size=10)
TITLE_FONT = Font(name="Consolas", bold=True, color=ORANGE, size=14)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="333333"),
)


def fetch_data(years: int = 10) -> pd.DataFrame:
    end = datetime.now()
    start = end - timedelta(days=years * 365)
    qqq = yf.download("QQQ", start=start, end=end, auto_adjust=True)
    tqqq = yf.download("TQQQ", start=start, end=end, auto_adjust=True)
    for d in [qqq, tqqq]:
        if isinstance(d.columns, pd.MultiIndex):
            d.columns = d.columns.get_level_values(0)
    df = pd.DataFrame(index=qqq.index)
    df["qqq_close"] = qqq["Close"]
    df["tqqq_close"] = tqqq["Close"]
    df = df.dropna()
    df["tqqq_return"] = df["tqqq_close"].pct_change().fillna(0)
    df["qqq_return"] = df["qqq_close"].pct_change().fillna(0)
    return df


def run_backtest(df, ema_period, sma_period):
    data = df.copy()
    data["ema"] = data["qqq_close"].ewm(span=ema_period, adjust=False).mean()
    data["sma"] = data["qqq_close"].rolling(window=sma_period).mean()
    warmup = max(ema_period, sma_period) + 5
    data = data.iloc[warmup:].copy()

    position = 0
    positions = []
    for i in range(len(data)):
        close = data["qqq_close"].iloc[i]
        ema = data["ema"].iloc[i]
        sma = data["sma"].iloc[i]
        if position == 0 and close >= ema:
            position = 1
        elif position == 1 and close < sma:
            position = 0
        positions.append(position)

    data["position"] = positions
    data["pos_shifted"] = data["position"].shift(1).fillna(0).astype(int)
    daily_cash = (1 + CASH_YIELD) ** (1 / 252) - 1
    data["strat_return"] = np.where(data["pos_shifted"] == 1, data["tqqq_return"], daily_cash)
    data["equity"] = INITIAL_CAPITAL * (1 + data["strat_return"]).cumprod()

    final = data["equity"].iloc[-1]
    years = len(data) / 252
    cagr = (final / INITIAL_CAPITAL) ** (1 / years) - 1
    peak = data["equity"].cummax()
    dd = (data["equity"] - peak) / peak
    max_dd = dd.min()
    std = data["strat_return"].std()
    sharpe = data["strat_return"].mean() / std * np.sqrt(252) if std > 0 else 0
    downside = data["strat_return"][data["strat_return"] < 0].std()
    sortino = data["strat_return"].mean() / downside * np.sqrt(252) if downside > 0 else 0
    calmar = cagr / abs(max_dd) if max_dd != 0 else 0
    time_in_market = data["pos_shifted"].mean()

    trade_changes = data["pos_shifted"].diff().fillna(0)
    num_trades = (trade_changes == 1).sum()

    entries = data[trade_changes == 1].index.tolist()
    exits = data[trade_changes == -1].index.tolist()
    trade_returns = []
    for i in range(min(len(entries), len(exits))):
        ei = data.index.get_loc(entries[i])
        xi = data.index.get_loc(exits[i])
        if xi > ei:
            trade_returns.append(data["equity"].iloc[xi] / data["equity"].iloc[ei] - 1)
    trade_wr = np.mean([r > 0 for r in trade_returns]) if trade_returns else 0

    data["year"] = data.index.year
    yearly = data.groupby("year")["strat_return"].apply(lambda x: (1 + x).prod() - 1).to_dict()

    return {
        "label": f"EMA{ema_period}/SMA{sma_period}",
        "ema": ema_period, "sma": sma_period,
        "cagr": cagr, "total_return": final / INITIAL_CAPITAL - 1,
        "max_dd": max_dd, "sharpe": sharpe, "sortino": sortino, "calmar": calmar,
        "num_trades": num_trades, "trade_wr": trade_wr,
        "time_in_market": time_in_market, "final_equity": final,
        "yearly": yearly,
    }


def compute_benchmark(df, col, label):
    data = df.copy()
    data["bh_return"] = data[col].pct_change().fillna(0)
    data["equity"] = INITIAL_CAPITAL * (1 + data["bh_return"]).cumprod()
    final = data["equity"].iloc[-1]
    years = len(data) / 252
    cagr = (final / INITIAL_CAPITAL) ** (1 / years) - 1
    peak = data["equity"].cummax()
    dd = (data["equity"] - peak) / peak
    max_dd = dd.min()
    std = data["bh_return"].std()
    sharpe = data["bh_return"].mean() / std * np.sqrt(252) if std > 0 else 0
    downside = data["bh_return"][data["bh_return"] < 0].std()
    sortino = data["bh_return"].mean() / downside * np.sqrt(252) if downside > 0 else 0
    calmar = cagr / abs(max_dd) if max_dd != 0 else 0
    data["year"] = data.index.year
    yearly = data.groupby("year")["bh_return"].apply(lambda x: (1 + x).prod() - 1).to_dict()
    return {
        "label": label, "ema": "-", "sma": "-",
        "cagr": cagr, "total_return": final / INITIAL_CAPITAL - 1,
        "max_dd": max_dd, "sharpe": sharpe, "sortino": sortino, "calmar": calmar,
        "num_trades": 1, "trade_wr": 1.0 if final > INITIAL_CAPITAL else 0.0,
        "time_in_market": 1.0, "final_equity": final, "yearly": yearly,
    }


def pct_font(val):
    return GREEN_FONT if val >= 0 else RED_FONT


def write_pct(ws, row, col, val):
    ws.cell(row=row, column=col, value=val)
    ws.cell(row=row, column=col).font = pct_font(val)
    ws.cell(row=row, column=col).number_format = '0.0%'
    ws.cell(row=row, column=col).fill = DARK_FILL
    ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")


def write_num(ws, row, col, val, fmt='0.00'):
    ws.cell(row=row, column=col, value=val)
    ws.cell(row=row, column=col).font = pct_font(val)
    ws.cell(row=row, column=col).number_format = fmt
    ws.cell(row=row, column=col).fill = DARK_FILL
    ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")


def write_text(ws, row, col, val, font=DATA_FONT, fill=DARK_FILL):
    ws.cell(row=row, column=col, value=val)
    ws.cell(row=row, column=col).font = font
    ws.cell(row=row, column=col).fill = fill
    ws.cell(row=row, column=col).border = THIN_BORDER


def write_header(ws, row, col, val):
    ws.cell(row=row, column=col, value=val)
    ws.cell(row=row, column=col).font = HEADER_FONT
    ws.cell(row=row, column=col).fill = HEADER_FILL
    ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")


def build_summary_sheet(wb, results, benchmarks):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = ORANGE

    headers = ["Combo", "CAGR", "Total Return", "Max DD", "Sharpe", "Sortino",
               "Calmar", "# Trades", "Win Rate", "In Market", "Final Equity"]

    # Title
    ws.merge_cells("A1:K1")
    ws.cell(row=1, column=1, value="TQQQ EMA/SMA Strategy — All Combinations (4.5% cash yield)")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = DARK_FILL

    # Headers
    for c, h in enumerate(headers, 1):
        write_header(ws, 3, c, h)

    row = 4
    # Benchmarks
    for b in benchmarks:
        write_text(ws, row, 1, b["label"], font=Font(name="Consolas", bold=True, color=ORANGE, size=10), fill=BENCHMARK_FILL)
        write_pct(ws, row, 2, b["cagr"])
        write_pct(ws, row, 3, b["total_return"])
        write_pct(ws, row, 4, b["max_dd"])
        write_num(ws, row, 5, b["sharpe"])
        write_num(ws, row, 6, b["sortino"])
        write_num(ws, row, 7, b["calmar"])
        write_num(ws, row, 8, b["num_trades"], fmt='0')
        write_pct(ws, row, 9, b["trade_wr"])
        write_pct(ws, row, 10, b["time_in_market"])
        write_num(ws, row, 11, b["final_equity"], fmt='$#,##0')
        for c in range(2, 12):
            ws.cell(row=row, column=c).fill = BENCHMARK_FILL
        row += 1

    row += 1  # gap

    # Strategy rows sorted by Calmar
    sorted_results = sorted(results, key=lambda x: x["calmar"], reverse=True)
    for r in sorted_results:
        write_text(ws, row, 1, r["label"])
        write_pct(ws, row, 2, r["cagr"])
        write_pct(ws, row, 3, r["total_return"])
        write_pct(ws, row, 4, r["max_dd"])
        write_num(ws, row, 5, r["sharpe"])
        write_num(ws, row, 6, r["sortino"])
        write_num(ws, row, 7, r["calmar"])
        write_num(ws, row, 8, r["num_trades"], fmt='0')
        write_pct(ws, row, 9, r["trade_wr"])
        write_pct(ws, row, 10, r["time_in_market"])
        write_num(ws, row, 11, r["final_equity"], fmt='$#,##0')

        # Highlight top calmar
        if r["calmar"] >= 0.70:
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = LIGHT_GREEN_FILL
        row += 1

    # Column widths
    widths = [15, 10, 13, 10, 10, 10, 10, 10, 10, 10, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Freeze panes
    ws.freeze_panes = "A4"


def build_yearly_sheet(wb, results, benchmarks):
    ws = wb.create_sheet("Yearly Returns")
    ws.sheet_properties.tabColor = BLUE

    all_items = benchmarks + sorted(results, key=lambda x: x["calmar"], reverse=True)
    all_years = sorted(set(y for item in all_items for y in item["yearly"].keys()))

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_years) + 1)
    ws.cell(row=1, column=1, value="Yearly Returns by Combo")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = DARK_FILL

    # Headers
    write_header(ws, 3, 1, "Combo")
    for c, year in enumerate(all_years, 2):
        write_header(ws, 3, c, str(year))

    row = 4
    for item in all_items:
        is_bm = item["label"] in ("TQQQ Buy & Hold", "QQQ Buy & Hold")
        fill = BENCHMARK_FILL if is_bm else DARK_FILL
        font = Font(name="Consolas", bold=True, color=ORANGE, size=10) if is_bm else DATA_FONT
        write_text(ws, row, 1, item["label"], font=font, fill=fill)

        for c, year in enumerate(all_years, 2):
            val = item["yearly"].get(year, None)
            if val is not None:
                ws.cell(row=row, column=c, value=val)
                ws.cell(row=row, column=c).font = pct_font(val)
                ws.cell(row=row, column=c).number_format = '0.0%'
                ws.cell(row=row, column=c).fill = fill
                # Conditional color fill for heatmap effect
                if not is_bm:
                    if val > 0.5:
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
                    elif val > 0.1:
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                    elif val > 0:
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="1A2E1A", end_color="1A2E1A", fill_type="solid")
                    elif val > -0.1:
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="2E1A1A", end_color="2E1A1A", fill_type="solid")
                    elif val > -0.3:
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="4A1010", end_color="4A1010", fill_type="solid")
                    else:
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 15
    for c in range(2, len(all_years) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = "B4"


def build_key_combos_sheet(wb, results):
    ws = wb.create_sheet("Key Combos")
    ws.sheet_properties.tabColor = GREEN

    key_labels = [
        "EMA30/SMA50", "EMA17/SMA23", "EMA30/SMA11", "EMA13/SMA20",
        "EMA30/SMA5", "EMA15/SMA20", "EMA20/SMA40", "EMA20/SMA50",
        "EMA17/SMA29", "EMA19/SMA23",
    ]

    key_combos = []
    for label in key_labels:
        match = [r for r in results if r["label"] == label]
        if match:
            key_combos.append(match[0])

    if not key_combos:
        return

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(key_combos) + 1)
    ws.cell(row=1, column=1, value="Key Combo Comparison")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = DARK_FILL

    # Headers
    write_header(ws, 3, 1, "Metric")
    for c, combo in enumerate(key_combos, 2):
        write_header(ws, 3, c, combo["label"])

    # Metrics
    metrics = [
        ("CAGR", "cagr", True),
        ("Total Return", "total_return", True),
        ("Max Drawdown", "max_dd", True),
        ("Sharpe", "sharpe", False),
        ("Sortino", "sortino", False),
        ("Calmar", "calmar", False),
        ("# Trades", "num_trades", False),
        ("Win Rate", "trade_wr", True),
        ("Time in Market", "time_in_market", True),
        ("Final Equity", "final_equity", False),
    ]

    row = 4
    for name, key, is_pct in metrics:
        write_text(ws, row, 1, name, font=Font(name="Consolas", bold=True, color=WHITE, size=10))
        for c, combo in enumerate(key_combos, 2):
            val = combo[key]
            if key == "final_equity":
                write_num(ws, row, c, val, fmt='$#,##0')
            elif key == "num_trades":
                write_num(ws, row, c, val, fmt='0')
            elif is_pct:
                write_pct(ws, row, c, val)
            else:
                write_num(ws, row, c, val)
        row += 1

    # Separator
    row += 1
    write_text(ws, row, 1, "YEARLY", font=Font(name="Consolas", bold=True, color=ORANGE, size=10))
    row += 1

    # Yearly breakdown
    all_years = sorted(set(y for c in key_combos for y in c["yearly"].keys()))
    for year in all_years:
        write_text(ws, row, 1, str(year))
        for c, combo in enumerate(key_combos, 2):
            val = combo["yearly"].get(year, 0)
            write_pct(ws, row, c, val)
        row += 1

    ws.column_dimensions["A"].width = 16
    for c in range(2, len(key_combos) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B4"


def build_consistency_sheet(wb, results):
    ws = wb.create_sheet("Consistency")
    ws.sheet_properties.tabColor = "9C27B0"

    all_years = sorted(set(y for r in results for y in r["yearly"].keys()))

    top10_count = {r["label"]: 0 for r in results}
    bottom5_count = {r["label"]: 0 for r in results}

    for year in all_years:
        year_returns = [(r["label"], r["yearly"].get(year, None)) for r in results]
        year_returns = [(l, v) for l, v in year_returns if v is not None]
        year_returns.sort(key=lambda x: x[1], reverse=True)
        for label, _ in year_returns[:10]:
            top10_count[label] += 1
        for label, _ in year_returns[-5:]:
            bottom5_count[label] += 1

    rows = []
    for r in results:
        rows.append({
            "label": r["label"],
            "top10": top10_count[r["label"]],
            "bottom5": bottom5_count[r["label"]],
            "calmar": r["calmar"],
            "cagr": r["cagr"],
            "max_dd": r["max_dd"],
        })
    rows.sort(key=lambda x: (-x["top10"], x["bottom5"]))

    # Title
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="Consistency Analysis — Top 10 / Bottom 5 Rankings per Year")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = DARK_FILL

    headers = ["Combo", f"Top 10 (of {len(all_years)}yr)", f"Bottom 5 (of {len(all_years)}yr)", "Calmar", "CAGR", "Max DD"]
    for c, h in enumerate(headers, 1):
        write_header(ws, 3, c, h)

    row = 4
    for r in rows:
        fill = DARK_FILL
        if r["top10"] >= 4 and r["bottom5"] == 0:
            fill = LIGHT_GREEN_FILL
        elif r["bottom5"] >= 3:
            fill = LIGHT_RED_FILL

        write_text(ws, row, 1, r["label"], fill=fill)
        write_num(ws, row, 2, r["top10"], fmt='0')
        ws.cell(row=row, column=2).fill = fill
        write_num(ws, row, 3, r["bottom5"], fmt='0')
        ws.cell(row=row, column=3).fill = fill
        write_num(ws, row, 4, r["calmar"])
        ws.cell(row=row, column=4).fill = fill
        write_pct(ws, row, 5, r["cagr"])
        ws.cell(row=row, column=5).fill = fill
        write_pct(ws, row, 6, r["max_dd"])
        ws.cell(row=row, column=6).fill = fill
        row += 1

    widths = [15, 18, 20, 10, 10, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"


def main():
    print("Fetching QQQ and TQQQ data (10 years)...")
    df = fetch_data(years=10)
    print(f"Got {len(df)} trading days ({df.index[0].date()} to {df.index[-1].date()})")

    print("Computing benchmarks...")
    benchmarks = [
        compute_benchmark(df, "tqqq_close", "TQQQ Buy & Hold"),
        compute_benchmark(df, "qqq_close", "QQQ Buy & Hold"),
    ]

    combos = list(product(ENTRY_EMAS, EXIT_SMAS))
    print(f"Running {len(combos)} parameter combinations...")
    results = []
    for ema, sma in combos:
        r = run_backtest(df, ema, sma)
        results.append(r)
    print(f"Completed {len(results)} backtests.")

    print("Building Excel report...")
    wb = Workbook()

    build_summary_sheet(wb, results, benchmarks)
    build_yearly_sheet(wb, results, benchmarks)
    build_key_combos_sheet(wb, results)
    build_consistency_sheet(wb, results)

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(REPORT_PATH))
    print(f"Report saved to: {REPORT_PATH.resolve()}")


if __name__ == "__main__":
    main()
